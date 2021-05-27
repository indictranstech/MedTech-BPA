from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe.utils import nowdate,nowtime, today, flt
from datetime import timedelta,date
import datetime
import calendar
import json
import time
from erpnext.stock.stock_ledger import get_previous_sle
import operator
import itertools

from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter
# reload(sys)
# sys.setdefaultencoding('utf8')

@frappe.whitelist()
def get_planing_master_details(filters=None):
	filters = json.loads(filters)
	
	data =[]
	from_date = datetime.date.today()
	precision=frappe.db.get_singles_value('System Settings', 'float_precision')
	planning_master=frappe.db.sql("""SELECT name, date(from_date) as from_date , date(to_date) as to_date ,description  from `tabPlanning Master` where {0} """.format(get_filters_codition(filters)), as_dict=1)
	pm_from_date = frappe.db.get_value('Planning Master', {'name' : planning_master[0].name}, 'from_date')
	pm_to_date = frappe.db.get_value('Planning Master', {'name' : planning_master[0].name}, 'to_date')
	bom_name = frappe.db.sql("""SELECT name, bom, amount from `tabPlanning Master Item` where planning_master_parent='{0}' and date>='{1}' and date<='{2}'""".format(planning_master[0].get('name'), planning_master[0].get('from_date'), planning_master[0].get('to_date')), as_dict=1)

	unic_bom = []
	bom_items=[]
	for row in bom_name:
		if row.bom not in unic_bom:
			unic_bom.append(row.bom)
		bom_data = frappe.db.sql("""SELECT a.name, a.quantity, b.item_code, b.item_name, b.stock_qty from `tabBOM` a join `tabBOM Explosion Item` b on a.name=b.parent where a.name='{0}'""".format(row.bom), as_dict=1)
		
		for bom in bom_data:
			bom['planning_master_item'] = row.name
			bom_items.append(bom)


	#planing_qty calculation from bom and planning master item
	planing_qty=0.0
	for plan_qty in bom_name:
		for row in bom_items:
			if row.planning_master_item==plan_qty.name:
				planing_qty=flt(row.stock_qty)/flt(row.quantity)*flt(plan_qty.amount)
				row["planing_qty"]=planing_qty

	#group_by planning qty base on item_code
	values_by_item = {}
	for d in bom_items:
		values_by_item.setdefault(d['item_code'], []).append(int(d['planing_qty']))
	new_data=[{'item_code':k, 'planing_qty':sum(v)} for k, v in values_by_item.items()]

	
	warehouse = frappe.db.sql("""SELECT warehouse from `tabRM Warehouse List` where parent='MedTech Settings'""", as_dict=1)
	
	posting_time = nowtime()
	for row in new_data:
		row['from_date'] = from_date
		row['to_date'] = (planning_master[0].get('to_date'))
		row['po_qty']=0.0
		row['consider_po_qty']=0.0
		row['current_stock'] = 0.0
		
		warehouse_list = []
		#current stock calculation of item
		for wh in warehouse:
			child_warehouse = frappe.db.get_descendants('Warehouse', wh.warehouse)
			
			if len(child_warehouse) > 0:
				for ware in child_warehouse:
					warehouse_list.append(ware)
					# last_entry = get_previous_sle({
					# 	"item_code": row.get('item_code'),
					# 	"warehouse" : ware,
					# 	"posting_date": from_date,
					# 	"posting_time": posting_time })
					# row['current_stock'] += last_entry.qty_after_transaction if last_entry else 0.0
			else:
				warehouse_list.append(wh.warehouse)
				# last_entry = get_previous_sle({
				# 		"item_code": row.get('item_code'),
				# 		"warehouse" :  wh.warehouse,
				# 		"posting_date": from_date,
				# 		"posting_time": posting_time })
				# row['current_stock'] += last_entry.qty_after_transaction if last_entry else 0.0
		if len(warehouse_list) > 0:
			ohs_query = frappe.db.sql("SELECT item.item_code, sum(IFNULL (bin.actual_qty,0.0)) as ohs from `tabItem` item LEFT JOIN `tabBin` bin on item.item_code = bin.item_code  and item.disabled = 0 and bin.warehouse in {0} group by item.item_code".format(tuple(warehouse_list)), as_dict=1)
			
		else:
			ohs_query = frappe.db.sql("SELECT item.item_code, sum(IFNULL (bin.actual_qty,0.0)) as ohs from `tabItem` item LEFT JOIN `tabBin` bin on item.item_code = bin.item_code  and item.disabled = 0 and bin.warehouse = {0} group by item.item_code".format(warehouse_list[0]), as_dict=1)
		ohs_detail = {row.item_code : row.ohs for row in ohs_query}
		row['current_stock'] += ohs_detail.get(row.get("item_code"))
		
	#calculate pending po qty
	po_data = []
	for row in new_data:
		po_details = frappe.db.sql("""SELECT a.name, a.supplier, b.item_code,(b.qty-b.received_qty) as qty from `tabPurchase Order` a join `tabPurchase Order Item` b on a.name=b.parent where a.docstatus=1  and b.item_code='{0}' and a.status != 'Closed' and a.status != 'On Hold' and (b.qty-b.received_qty) > 0 """.format(row.get('item_code')), as_dict=1)

		for po in po_details:
			# accept_qty = frappe.db.get_values("Purchase Receipt Item", {'purchase_order':po.get('name'), 'item_code':po.get('item_code')}, ['item_code', 'actual_accepted_qty', 'parent', 'purchase_order'], as_dict=1)
			# if accept_qty and po.get('item_code') == accept_qty[0].get('item_code'):
			po['qty'] = po.get('qty')
			po_data.append(po)
		
	#calculate pending po qty for supplier
	supplier = []
	for row in new_data:			
		for poqty in po_data:
			if row.get('item_code') == poqty.item_code:
				row['po_qty']+=flt(poqty.qty, precision)
				if row['po_qty'] > 0 :
					supplier.append({'supplier':poqty.supplier, 'qty':poqty.qty, 'item_code':poqty.item_code})
	
	outputList=[]
	for i,g in itertools.groupby(supplier, key=operator.itemgetter("item_code")):
		outputList.append(list(g))
	
	for row in new_data:
		for out in outputList:
			if row.get('item_code') == out[0].get('item_code'):
				values_by_supplier = {}
				for d in out:
					values_by_supplier.setdefault(d['supplier'], []).append(int(d['qty']))
				supplier_data=[{'supplier':k, 'qty':sum(v)} for k, v in values_by_supplier.items()]
				row['supplier']=supplier_data

		row['consider_po_qty']=flt((row.get('current_stock')+row.get('po_qty'))-row.get('planing_qty'), precision)
		row['no_consider_po_qty']=flt((row.get('current_stock')-row.get('planing_qty')),precision)


	#Production plan qty calculation
	pp_item = []
	for row in unic_bom:
		pp_data = frappe.db.sql("""SELECT c.item_code, c.quantity from `tabProduction Plan` a join `tabProduction Plan Item` b on a.name=b.parent join `tabMaterial Request Plan Item` c on a.name=c.parent where b.bom_no='{0}' and a.docstatus=1 and (a.status='Submitted' or a.status='In Process') and a.posting_date>'{1}' and posting_date<'{2}'""".format(row, planning_master[0].get('from_date'), from_date), as_dict=1)
		for row in pp_data:
			pp_item.append(row)

	#group_by pp qty base on item_code
	values_by_item = {}
	for d in pp_item:
		values_by_item.setdefault(d['item_code'], []).append(int(d['quantity']))
	pp_data=[{'item_code':k, 'qty':sum(v)} for k, v in values_by_item.items()]

	for row in new_data:
		row['item_name']=frappe.db.get_value("Item", {'item_code':row.get('item_code')}, 'item_name')
		for pp in pp_data:
			if row.get('item_code') == pp.get('item_code'):
				row['consider_po_qty'] = flt((row.get('consider_po_qty')-pp.get('qty')),precision)
				row['no_consider_po_qty'] = flt((row.get('no_consider_po_qty')-pp.get('qty')),precision)

	
	planning_data = dict()
	planning_data['from_date'] = (pm_from_date).strftime('%d-%m-%Y') if pm_from_date else ''
	
	planning_data['to_date'] = (pm_to_date).strftime('%d-%m-%Y') if pm_to_date else ''
	planning_data['description'] = planning_master[0].description if planning_master[0].description else ''
	planning_data['planning_master'] = planning_master[0].name
	new_data[0]['planning_data'] = planning_data
	path = 'medtech_bpa/medtech_bpa/page/supplier_wise_rm_wis/supplier_wise_rm_wis.html'
	html=frappe.render_template(path,{'data':new_data})
	return {'html':html, 'data':new_data}


def get_filters_codition(filters):
	conditions = "1=1"
	if filters.get("planning_master"):
		conditions += " and name = '{0}'".format(filters.get('planning_master'))
	return conditions

@frappe.whitelist()
def custome_report_to_pdf(html, orientation="Landscape"):
	frappe.local.response.filename = "report.pdf"
	frappe.local.response.filecontent = get_pdf(html, {"orientation": orientation,'page-size':'A4'})
	frappe.local.response.type = "download"

@frappe.whitelist()
def make_xlsx_file(renderd_data):
	data =json.loads(renderd_data)

	
	from_date = data[0].get('from_date')
	to_date = data[0].get('to_date')

	book = Workbook()
	sheet = book.active
	
	row = 1
	col = 1

	cell = sheet.cell(row=row,column=col)
	cell.value = 'Planning Master'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+1)
	cell.value = data[0].get("planning_data").get("planning_master")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+3)
	cell.value = 'From Date'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+4)
	cell.value =data[0].get("planning_data").get('from_date')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	cell = sheet.cell(row=row,column=col+6)
	cell.value = 'To Date'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+7)
	cell.value = data[0].get("planning_data").get('to_date')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	cell = sheet.cell(row=row,column=col+9)
	cell.value = 'Description'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+10)
	cell.value = data[0].get("planning_data").get('description')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	row = 2
	col = 1
	
	cell = sheet.cell(row=row,column=col)
	cell.value = 'RM & Supplier wise Report'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)


	row+=1
	col = 1
	cell = sheet.cell(row=row,column=col)
	cell.value = 'Select Period for Planning'
	cell.font = cell.font.copy(bold=True, center=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)

	row=4
	col = 1
	cell = sheet.cell(row=row,column=col)
	cell.value = 'From Date:{0}'.format(from_date)
	cell.font = cell.font.copy(bold=True)
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)

	col = 4
	cell = sheet.cell(row=row,column=col)
	cell.value = 'To Date'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=4, start_column=4, end_row=4, end_column=4)

	col = 5
	cell = sheet.cell(row=row,column=col)
	cell.value =to_date
	cell.font = cell.font.copy(bold=True)
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=4, start_column=5, end_row=4, end_column=6)

	row = 5
	col = 1
	cell = sheet.cell(row=row,column=col)
	cell.value ='Item Name'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=5, start_column=1, end_row=6, end_column=1)

	col = 2
	cell = sheet.cell(row=row,column=col)
	cell.value ='Total Pending PO'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)

	col = 3
	cell = sheet.cell(row=row,column=col)
	cell.value ='Shortage/ Excess Qty'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=5, start_column=3, end_row=5, end_column=4)

	col = 5
	cell = sheet.cell(row=row,column=col)
	cell.value ='Supplier'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=5, start_column=5, end_row=6, end_column=5)

	col = 6
	cell = sheet.cell(row=row,column=col)
	cell.value ='PO Qty'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=5, start_column=6, end_row=6, end_column=6)


	row = 6
	col = 3
	cell = sheet.cell(row=row,column=col)
	cell.value ='Considered PO'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=6, start_column=3, end_row=6, end_column=3)


	col = 4
	cell = sheet.cell(row=row,column=col)
	cell.value ='Not Considered PO'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
	sheet.merge_cells(start_row=6, start_column=4, end_row=6, end_column=4)

	row+=1
	col = 1
	for d in data:
		cell = sheet.cell(row=row,column=col)
		cell.value = d.get("item_code")+':'+d.get("item_name")
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell = sheet.cell(row=row,column=col+1)
		cell.value = d.get("po_qty")
		if d.get("po_qty") < 0 :
			cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell = sheet.cell(row=row,column=col+2)
		cell.value = d.get("consider_po_qty")
		if d.get("consider_po_qty") < 0:
			cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell = sheet.cell(row=row,column=col+3)
		cell.value = d.get("no_consider_po_qty")
		if d.get("no_consider_po_qty"):
			cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		# row+=1
		supplier_data=d.get('supplier')
		if supplier_data:
			for supplier in supplier_data:
				cell = sheet.cell(row=row,column=col)
				cell.value = d.get('item_name')
				cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
				

				cell = sheet.cell(row=row,column=col+4)
				cell.value = supplier.get('supplier')
				cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
				cell = sheet.cell(row=row,column=col+5)
				cell.value = supplier.get('qty')
				if supplier.get("qty") < 0 :
					cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
				cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
				row+=1
		else:
			cell = sheet.cell(row=row,column=col+4)
			cell.value = ''
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell = sheet.cell(row=row,column=col+5)
			cell.value = ''
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			row+=1


	file_path = frappe.utils.get_site_path("public")
	book.save(file_path+'/supplier_wise_rm_wise.xlsx')

@frappe.whitelist()
def download_xlsx():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	wb = openpyxl.load_workbook(filename=file_path+'/supplier_wise_rm_wise.xlsx')
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	frappe.local.response.filename = "supplier_wise_rm_wise.xlsx"
