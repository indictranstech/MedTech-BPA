from __future__ import unicode_literals
import frappe
from frappe.model.document import Document
from frappe.utils import nowdate, cstr, flt, cint, now, getdate,get_datetime,time_diff_in_seconds,add_to_date,time_diff_in_seconds,add_days,today
from datetime import datetime,date, timedelta
from frappe.model.naming import make_autoname
import time
import json
import pandas as pd
from frappe import _
from erpnext.manufacturing.doctype.bom.bom import get_bom_items_as_dict
import operator
import itertools

from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter

@frappe.whitelist()
def get_planning_master_data(filters=None):
	filters = json.loads(filters)
	data = dict()
	precision=frappe.db.get_singles_value('System Settings', 'float_precision')
	planning_master = filters.get("planning_master")
	pm_from_date = frappe.db.get_value('Planning Master', {'name' : planning_master}, 'from_date')
	pm_to_date = frappe.db.get_value('Planning Master', {'name' : planning_master}, 'to_date')
	pm_description = frappe.db.get_value('Planning Master', {'name' : planning_master}, 'description')
	# fetch date_data 
	date_data = get_date_data(planning_master)

	data.update({'date_data':[(pmi.get('date')).strftime('%d-%m-%Y') for pmi in date_data]})


	# fetch date_data_with_amount
	date_wise_planning_details = date_wise_planning_master(planning_master)

	data.update({'date_data_with_amount':date_wise_planning_details})

	company = frappe.db.get_single_value("Global Defaults",'default_company')
	# fetch warehouse list from medtech settings
	fg_warehouse_list = get_warehouses()
	ohs_dict = get_current_stock(fg_warehouse_list)
	

	# fetch planning_data from planning_master item
	planning_data = get_planning_data(planning_master)

	for row in date_wise_planning_details:
		for item in planning_data:
			if row.item_code == item.item_code:
				item.update({row.get('date').strftime('%d-%m-%Y'):row.amount})


	item_dict = dict()
	final_dict = dict()

	# fetch raw materials data from BOM and stock data
	for item in planning_data:
		bom_data = get_bom_data(item.get('bom'),item.get("include_exploded_bom"))
		item.update({'bom_data':bom_data})

	item_details = {item.item_code : item for item in planning_data}
	remaining_dict = dict()
	check_date = dict()
	for date in date_wise_planning_details:
		fg_detail = item_details.get(date.get('item_code'))

		for raw in fg_detail.get('bom_data'):
			required_qty = raw.get('stock_qty') * date.get('amount')

			if raw.get('item_code') not in remaining_dict:
				current_stock = ohs_dict.get(raw.get('item_code')) or 0
				stock_expected = get_expected_stock(raw.get('item_code'), date.get('date'))
				stock_exp = stock_expected[0].get("qty") if stock_expected else 0
				virtual_stock = current_stock + stock_exp

				remaining_qty = virtual_stock - required_qty

				remaining_dict[raw.get('item_code')] = remaining_qty if remaining_qty > 0 else 0
				raw[date.get('date').strftime('%d-%m-%Y')] = flt(remaining_qty, precision) if remaining_qty < 0 else flt(required_qty, precision)
				check_date[raw.get('item_code')] = date.get('date')
			else:
				if date.get('date') == check_date.get(raw.get('item_code')):
					virtual_stock = remaining_dict.get(raw.get('item_code'))
				else:
					stock_expected = get_expected_stock(raw.get('item_code'), date.get('date'))
					stock_exp = stock_expected[0].get("qty") if stock_expected else 0
					check_date[raw.get('item_code')] = date.get('date')
					virtual_stock = stock_exp + remaining_dict.get(raw.get('item_code'))

				remaining_qty = virtual_stock - required_qty

				remaining_dict[raw.get('item_code')] = remaining_qty if remaining_qty > 0 else 0
				raw[date.get('date').strftime('%d-%m-%Y')] = flt(remaining_qty, precision) if remaining_qty < 0 else flt(required_qty, precision)

	data.update({'planning_data':item_details})
	data['planning_master'] = planning_master if planning_master else ''
	data['from_date'] = (pm_from_date).strftime('%d-%m-%Y') if pm_from_date else ''
	data['to_date'] = (pm_to_date).strftime('%d-%m-%Y') if pm_to_date else ''
	data['description'] = pm_description if pm_description else ''
	final_data = data
	path = 'medtech_bpa/medtech_bpa/page/plan_availability/plan_availability.html'
	html=frappe.render_template(path,{'data':data})
	return {'html':html,'data':final_data}


def get_filters_codition(filters):
	conditions = "1=1"
	if filters.get("planning_master"):
		conditions += " and planning_master_parent = '{0}'".format(filters.get('planning_master'))
	return conditions

# fetch planning dates
def get_date_data(planning_master):
	date_data = frappe.db.sql(""" SELECT DISTINCT a.date from `tabPlanning Master Item` a join `tabPlanning Master` b on a.planning_master_parent = b.name  where a.planning_master_parent= '{0}' group by a.date order by a.date""".format(planning_master), as_dict=1)
	# date_data = date.strftime('%d-%m-%Y')
	return date_data

# fetch planning_qty and dates
def date_wise_planning_master(planning_master):
	date_data_with_amount = frappe.db.sql(""" SELECT a.item_code, a.date, a.amount from `tabPlanning Master Item` a join `tabPlanning Master` b on a.planning_master_parent = b.name  where a.planning_master_parent= '{0}' order by a.date asc""".format(planning_master), as_dict=1)
	return date_data_with_amount

# fetch FG data and it's details
def get_planning_data(planning_master):
	planning_data = frappe.db.sql(""" SELECT DISTINCT a.item_code,a.item_name, a.uom, sum(a.amount) as amount ,a.bom ,a.include_exploded_bom from `tabPlanning Master Item` a join `tabPlanning Master` b on a.planning_master_parent = b.name  where a.planning_master_parent= '{0}' group by a.item_code """.format(planning_master), as_dict=1)
	return planning_data

# fetch warehouse from medtech settings
def get_warehouses():
	# warehouse list
	fg_warehouse = frappe.db.sql("select warehouse from `tabRM Warehouse List`", as_dict = 1)
	from_warehouses = []

	if fg_warehouse:
		# fg_warehouse_ll = ["'" + row.warehouse + "'" for row in fg_warehouse]
		# fg_warehouse_list = ','.join(fg_warehouse_ll)
		for row in fg_warehouse:
			warehouse_list = frappe.db.get_descendants('Warehouse', row.warehouse)
			if warehouse_list:
				for item in warehouse_list:
					from_warehouses.append(item)
			else:
				from_warehouses.append(row.warehouse)
		fg_warehouse_ll = ["'" + row + "'" for row in from_warehouses]
		fg_warehouse_list = ','.join(fg_warehouse_ll)
	else:
	    fg_warehouse_list = "' '"
	
	# fg_warehouse = frappe.db.sql("SELECT warehouse from `tabFG Warehouse Group`", as_dict = 1)
	# fg_warehouse_list = tuple([item.warehouse for item in fg_warehouse])
	return fg_warehouse_list

def get_bom_data(bom,include_exploded_bom):
	if include_exploded_bom == 0:
		bom_data = frappe.db.sql("""SELECT b.item_code,b.stock_uom,b.qty as stock_qty from `tabBOM` a join `tabBOM Item` b on b.parent = a.name where a.name = '{0}'""".format(bom),as_dict=1)
	else:
		bom_data = frappe.db.sql("""SELECT b.item_code,b.stock_uom,b.stock_qty  from `tabBOM` a join `tabBOM Explosion Item` b on b.parent = a.name where a.name = '{0}'""".format(bom),as_dict=1)
	return bom_data

def get_available_item_qty(item_code, warehouses, company):
	# gets all items available in different warehouses
	warehouses = [x.get('name') for x in frappe.get_list("Warehouse", {'company': company}, "name")]

	filters = frappe._dict({
		'item_code': item_code,
		'warehouse': ['in', warehouses],
		'actual_qty': ['>=', 0]
	})

	item_locations = frappe.get_all('Bin',
		fields=['warehouse', 'actual_qty as qty'],
		filters=filters,
		# limit=required_qty,
		order_by='creation')

	return item_locations

def get_current_stock(fg_warehouse_list):
	current_stock = frappe.db.sql("""SELECT item_code,sum(actual_qty) as qty from `tabBin` where warehouse in ({0}) group by item_code """.format(fg_warehouse_list),as_dict=1,debug=1)
	ohs_dict = {item.item_code : item.qty for item in current_stock}
	return ohs_dict


def get_expected_stock(item,date):
	stock_expected = frappe.db.sql("""SELECT i.item_code, ifnull(sum((i.qty-i.received_qty)),0) as qty from `tabPurchase Order Item` i join `tabPurchase Order` p on p.name = i.parent where i.item_code = '{0}' and i.expected_delivery_date between '{1}' and '{2}' and p.status != 'On Hold' and  p.status != 'Closed' """.format(item,nowdate(),date),as_dict=1)
	return stock_expected



@frappe.whitelist()
def get_planning_dates(planning_master):
	planning_dates = frappe.db.sql("""SELECT from_date ,to_date ,description from `tabPlanning Master` where name = '{0}'""".format(planning_master),as_dict=1)
	date_dict = dict()
	date_dict['from_date'] =planning_dates[0].get('from_date').strftime('%d-%m-%Y')
	date_dict['to_date'] = planning_dates[0].get("to_date").strftime('%d-%m-%Y') 
	date_dict['description'] = planning_dates[0].get("description")
	return date_dict

@frappe.whitelist()
def make_xlsx_file(renderd_data):
	data =json.loads(renderd_data)
	
	header = ['Sr.No','FG Items','FG Item Name','UOM','FG Qty','BOM Items','BOM UOM']
	
	book = Workbook()
	sheet = book.active
	
	row = 1
	col = 1

	cell = sheet.cell(row=row,column=col)
	cell.value = 'Planning Master'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+1)
	cell.value = data.get("planning_master")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+3)
	cell.value = 'From Date'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+4)
	cell.value = data.get('from_date')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	cell = sheet.cell(row=row,column=col+6)
	cell.value = 'To Date'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+7)
	cell.value = data.get('to_date')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	cell = sheet.cell(row=row,column=col+9)
	cell.value = 'Description'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+10)
	cell.value = data.get('description')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	row = 2
	col = 1

	for item in header:
		cell = sheet.cell(row=row,column=col)
		cell.value = item
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
		
		col+=1


	col = 8
	for date in data.get("date_data"):
		cell = sheet.cell(row=2,column=col)
		cell.value = date
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
		col+=1
	
	row = 3 
	col  = 1
	count = 0
	
	for item in data.get("planning_data"):
		item_dict = data.get("planning_data").get(item)
		cell = sheet.cell(row=row,column=col)
		cell.value = count + 1
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+1)
		cell.value = item_dict.get("item_code") 
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+2)
		cell.value = item_dict.get("item_name")
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+3)
		cell.value = item_dict.get("uom")
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+4)
		cell.value = item_dict.get("amount")
		if item_dict.get("amount") < 0 :
			cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+5)
		cell.value = ''
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+6)
		cell.value = ''
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)


		col_date = 8
		for date in data.get("date_data"):
			cell = sheet.cell(row=row,column=col_date)
			cell.value = item_dict.get(date)
			if item_dict.get(date) < 0:
				cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)
			col_date+=1

		row+= 1
		bom_col = 6
		for bom_item in item_dict.get("bom_data"):
			cell = sheet.cell(row=row,column=bom_col)
			cell.value = bom_item.get("item_code")
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=bom_col, end_row=row, end_column=bom_col)

			cell = sheet.cell(row=row,column=bom_col+1)
			cell.value = bom_item.get("stock_uom")
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=bom_col + 1, end_row=row, end_column=bom_col+1)

			date_col = bom_col + 2 
			for date in data.get("date_data"):
				cell = sheet.cell(row=row,column=date_col)
				cell.value = bom_item.get(date)
				if bom_item.get(date) < 0 :
					cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
				cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
				sheet.merge_cells(start_row=row, start_column=date_col, end_row=row, end_column=date_col)
				date_col+=1
			row+=1

		count+=1

	file_path = frappe.utils.get_site_path("public")
	fname = "MRP_FG_WISE_REPORT" + nowdate() + ".xlsx"
	book.save(file_path+fname)

@frappe.whitelist()
def download_xlsx():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	now = datetime.now()
	fname = "MRP_FG_WISE_REPORT" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname

