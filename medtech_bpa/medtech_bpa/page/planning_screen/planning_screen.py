#Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt
from __future__ import unicode_literals
import frappe

from frappe.model.document import Document
from frappe.utils import nowdate, cstr, flt, cint, now, getdate,get_datetime,time_diff_in_seconds,add_to_date,time_diff_in_seconds,add_days,today
from datetime import datetime,date, timedelta
from frappe.model.naming import make_autoname,get_default_naming_series,parse_naming_series
import time
import json
import pandas as pd
from frappe import _
from frappe.utils import now_datetime

# from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter

# import xlsxwriter
# import csv
# import openpyxl
# from openpyxl import load_workbook
# from xlsxwriter import Workbook
import io
# import numpy as np
# from io import BytesIO

@frappe.whitelist()
def send_naming_series():
    today = now_datetime()
    n = ''
    naming_series = get_default_naming_series("Planning Master")
    parts = naming_series.split('.')[:-1]
    for e in parts:
      part = ''
      if e == 'YY':
        part = today.strftime('%y')
      elif e == 'MM':
        part = today.strftime('%m')
      elif e == 'DD':
        part = today.strftime("%d")
      elif e == 'YYYY':
        part = today.strftime('%Y')
      elif e == 'FY':
        part = frappe.defaults.get_user_default("fiscal_year")
      else:
        part = e
      n += part
    #naming_series = parse_naming_series(naming_series)
    return n  
@frappe.whitelist()
def save_items_data(description,from_date,to_date,data):
    
    dates_check(from_date,to_date)
    date_range = pd.date_range(from_date,to_date)
    date_range= [(i).strftime("%d-%m-%Y") for i in  date_range]
    doc = frappe.new_doc("Planning Master")
    doc.from_date=from_date
    doc.to_date=to_date
    data = json.loads(data)
   # doc.title=title
    doc.description=description
    #doc.name=title
    doc.save()
    #frappe.rename_doc("Planning Master",doc.name,doc.title,force=True)
    #frappe.db.set_value("Planning Master", doc.name,'title',doc.name)
    for j in date_range:
        for i in range(len(data['item_code'])):
            doc_child= frappe.new_doc('Planning Master Item')
            doc_child.item_code = data['item_code'][i]
            doc_child.item_name = data['item_name'][i]
            doc_child.date =  datetime.strptime(j,"%d-%m-%Y")
            try:
                if '<br>' in data[j[:-5]][i]:
                    doc_child.amount = float(data[j[:-5]][i][0:-4]) if data[j[:-5]][i][0:-4] not in ['',None,'.'] else 0
                else:
                    doc_child.amount =  float(data[j[:-5]][i]) 
            except Exception as e:
                frappe.throw("Error in saving data at column %s and row %s"%(j,i+1))
            doc_child.uom = data["uom"][i]
            doc_child.bom = data["bom"][i]
            doc_child.planning_master_parent = doc.name
            doc_child.save()
            frappe.db.set_value("Planning Master Item", doc_child.name,'title',doc_child.name)

    return ["1",doc.name] 

def dates_check(from_date,to_date):    
    if from_date <= today():
        frappe.throw("From Date must be greater than today.")
    if to_date <= today():
        frappe.throw("To Date must be greater than today.")
    if to_date < from_date:
        frappe.throw("To Date must be greater than or equals to From Date.")
   # date_overlap = frappe.db.sql("select name,from_date,to_date from `tabPlanning Master` where '%s' between from_date and to_date or '%s' between from_date and to_date"%(from_date,to_date),as_list=1)
    
    #if len(date_overlap)>0:
      #  frappe.throw("Date overlaps with Planning Master %s having From Date %s and To Date %s "%(date_overlap[0][0],date_overlap[0][1].strftime('%d-%m-%Y'),date_overlap[0][2].strftime('%d-%m-%Y')))

@frappe.whitelist()
def delete_data(delete_data):
    doc = frappe.get_doc("Planning Master",delete_data)
    if doc.from_date <= date.today():
        frappe.throw("The entry cannot be deleted because it contains past date or present date.")
    #names_of_child = [i[0] for i in frappe.db.sql("""select name from `tabPlanning Master Item` where planning_master_parent = '%s'"""%(delete_data))]
    #for i in names_of_child:
        #frappe.delete_doc("Planning Master Item",i)
    frappe.db.sql("""delete  from `tabPlanning Master Item` where planning_master_parent = '%s'"""%(delete_data))
    frappe.delete_doc("Planning Master",delete_data)
    frappe.db.commit()
    return "1"

@frappe.whitelist()
def get_items_data(from_date, to_date,listview=None):
        dates_check(from_date,to_date)    
        sdate =  datetime.strptime(from_date, '%Y-%m-%d').date()
        edate =  datetime.strptime(to_date, '%Y-%m-%d').date()
        delta = edate - sdate
        date_list = []
        date_dict= []
        for i in range(delta.days + 1):
            day = sdate + timedelta(days=i)
        
            if date.today() >= day:
                date_dict.append([day.strftime('%d-%m'),0])
            else:
                date_dict.append([day.strftime('%d-%m'),1])
            date_list.append(day.strftime('%d-%m-%Y'))
    
        data = dict()
        data['header_list'] = date_dict

        ig_list = []
        fg_item_group = frappe.db.sql("select item_group from `tabFG Item Group`", as_dict = 1)
        for ig in fg_item_group:
            child_ig = frappe.db.get_descendants('Item Group', ig.item_group)
            if child_ig:
                for row in child_ig:
                    ig_list.append(row)
            else:
                ig_list.append(ig.item_group)
        if fg_item_group:
            fg_group_list = ["'" + row.item_group + "'" for row in fg_item_group]
            fg_item_group_list = ','.join(fg_group_list)
        else:
            fg_item_group_list = "' '"

        item_detail = frappe.db.sql("select i.item_code,i.item_name, i.item_group, i.stock_uom from `tabItem` i join `tabBOM` b on i.item_code = b.item where b.docstatus = 1 and i.item_group in {0} group by i.item_code order by i.item_group, i.item_code".format(tuple(ig_list)), as_dict=1)
        # bom_query = frappe.db.sql("select b.item, b.name  from `tabBOM` b where b.docstatus = 1 and b.is_default = (select max(b2.is_default) from `tabBOM` b2 where b2.item = b.item and b.docstatus = 1)", as_dict = 1)

        bom_query = frappe.db.sql("select b.item, b.name  from `tabBOM` b where b.docstatus = 1 and b.is_default = 1 and b.is_active = 1", as_dict = 1)
        bom_dict = {bom.get('item') : bom.get('name') for bom in bom_query}
        uom_dict={}
        
        for i in item_detail:
            uom_dict[i.item_code]=[i.item_name,i.stock_uom]
       
        for row in item_detail:
            row['bom'] = bom_dict.get(row.get('item_code'))
        item_code_list = [i.get('item_code') for i in item_detail]
        data['item_data'] = item_detail
        data['item_code'] = item_code_list
        data['update']=0
        data['uom_dict']=uom_dict
       
        return data

@frappe.whitelist()
def get_bom_based_on_item_code(item_code):
   
    bom_list = [ i[0] for i in frappe.db.sql("select b.name from `tabItem` i join `tabBOM` b on i.item_code = b.item where i.item_code='%s'"%(item_code), as_list=1)]
    
    return bom_list
@frappe.whitelist()
def fetch_data(name):
    try:
        send_data={}
        date_range = pd.date_range(frappe.get_value('Planning Master',name,'from_date'),frappe.get_value('Planning Master',name,'to_date'))
        date_range= [[(i).strftime("%d-%m-%Y"),0 if pd.Timestamp(date.today()) >= i else 1] for i in  date_range]
        data = frappe.db.sql("""select GROUP_CONCAT(name) as name,item_code ,item_name, uom as stock_uom,bom ,GROUP_CONCAT(amount) as amount from `tabPlanning Master Item` where planning_master_parent = '%s' group by bom  order by creation"""%(name),as_dict=1)
        for i in data:
            i['name'] = i['name'].split(',')
            i['amount']= [float(i) for i in i['amount'].split(',')]
        send_data['header_list']=date_range
        send_data['item_data']=data
        send_data['update']=1
        return send_data
    except Exception as e:
        error_log= frappe.log_error(message=frappe.get_traceback(), title=str(e))
        frappe.throw("Error in fetching data {}".format(frappe.utils.get_link_to_form("Error Log",error_log.name)))

@frappe.whitelist()
def update_data(update_data):
    try :

        update_data = json.loads(update_data)
        if  len(update_data) ==0:
            return ["0"]
        #update_data = json.loads(update_data)
        for i in update_data:
            for values in range(len(i)):
                if type(i[values]) == str:
                    if '<br>' in i[values]:
                        i[values]=i[values][0:-4]
                        i[values] = 0 if i[values] == '' else i[values]
                    if '.' == i[values]:
                        i[values]=0
            doc = frappe.get_doc("Planning Master Item" ,i[0])
            #if doc.amount== int(i[2]):
            try:
                doc.amount=float(i[1])
            except Exception as e:
                return ["0","Please check if the values changed  are correct"]
            doc.save()
            #else:
                #frappe.throw("The old amount does not match for planning master item %s and amount %s"%(update_data[1],int(i[2])))
        return ["1",doc.planning_master_parent]
    except Exception as e:
        error_log= frappe.log_error(message=frappe.get_traceback(), title=str(e))
        frappe.throw("Error in updating data {}".format(frappe.utils.get_link_to_form("Error Log",error_log.name)))
        
        
@frappe.whitelist()
def return_list(item_detail):
    item_detail = [i[0] for i in frappe.db.sql("select b.name from `tabItem` i join `tabBOM` b on i.item_code = b.item where i.item_code='%s'"%(item_detail), as_list=1)]
    
    return item_detail





@frappe.whitelist()
def create_file(name= ""):

    data = fetch_data(name)

    now = datetime.now()
    dt_string = now.strftime("%d-%m-%Y_%H:%M:%S")

    file = str(time.time())
    # now = datetime.now()
    date_time = now.strftime("%m-%d-%Y")
    fname = "Planning Master_" +dt_string+ ".xlsx"
    # fname = "Planning Master_" +date_time+"_"+ now.strftime("%H:%M:%S") + ".xlsx"
    file_name = make_xlsx_csv(data, fname)
    return file_name

def make_xlsx_csv(data, fname):
    # Create a workbook and add a worksheet.
    # heading_col1 = {'subject': 'Item_Code'}
    # heading_col2 = {'subject': 'Item_Name   '}
    # heading_col3 = {'subject': 'UOM '}
    # heading_col4 = {'subject': 'BOM'}

    file = frappe.utils.get_site_path("public")+"/"+ fname
    # workbook = Workbook(file)
    # worksheet = workbook.add_worksheet()
    # bold = workbook.add_format({'bold': True, 'bg_color': '#5e64ff','border':1,'border_color':'#000000','align':'center','font_color': 'fdfbfb'})
    # bold.set_align('vcenter')
    # bold.set_font_name('Times New Roman')
    # bold1 = workbook.add_format({'border':1,'border_color':'#000000','align':'center'})
    # worksheet.set_column('A:BE', 18)

    # # header_list = data.get('header_list')
    # header_list = []
    # header_list.insert(0, heading_col1)
    # header_list.insert(1, heading_col2)
    # header_list.insert(2, heading_col3)
    # header_list.insert(3, heading_col4)

    # init_cnt = 4
    # for row in data.get('header_list'):
    #     temp_dict = {'subject':row[0] }
    #     header_list.insert(init_cnt, temp_dict)
    #     init_cnt = init_cnt + 1

    # cnt = 0
    # for col in header_list:
    #     worksheet.write(1, cnt, col['subject'], bold)
    #     cnt = cnt + 1

    # final_data_list = []
    # for modify_data in data.get("item_data"):
    #     temp_list = []
    #     temp_list.append(modify_data["item_code"])
    #     temp_list.append(modify_data["item_name"])
    #     temp_list.append(modify_data["stock_uom"])
    #     temp_list.append(modify_data["bom"])
    #     for date_data in modify_data.get("amount"):
    #         temp_list.append(date_data)
    #     final_data_list.append(temp_list)

    # row_cnt = 2
    # col_cnt = 0
    # index_list = [0,1,2,3]
    # for details in final_data_list:
    #     col_cnt = 0
    #     for value in details:
    #         index = details.index(value)
    #         if (index in index_list):
    #             cell_format = workbook.add_format({'border':1,'border_color':'#000000','align':'left'})
    #             cell_format.set_align('vcenter')
    #             cell_format.set_font_name('Times New Roman')
    #         else:
    #             cell_format = workbook.add_format({'border':1,'border_color':'#000000','align':'right'})
    #             cell_format.set_align('vcenter')
    #             cell_format.set_font_name('Times New Roman')

    #         worksheet.write(row_cnt, col_cnt, value,cell_format)
    #         col_cnt += 1
    #     row_cnt += 1

    # workbook.close()
    
    
    book = Workbook()
    sheet = book.active


    heading_col1 = {'subject': 'Item_Code'}
    heading_col2 = {'subject': 'Item_Name'}
    heading_col3 = {'subject': 'UOM'}
    heading_col4 = {'subject': 'BOM'}

    header_list = []
    header_list.insert(0, heading_col1)
    header_list.insert(1, heading_col2)
    header_list.insert(2, heading_col3)
    header_list.insert(3, heading_col4)

    init_cnt = 4
    for row in data.get('header_list'):
        temp_dict = {'subject':row[0] }
        header_list.insert(init_cnt, temp_dict)
        init_cnt = init_cnt + 1

    column_width = ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z','AA','AB','AC','AD','AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ','BA','BB','BC','BD','BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ']
    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['D'].width = 18
    for i in column_width:
        sheet.column_dimensions[i].width = 15

    
    row_cnt = 1
    col_cnt = 1

    for item in header_list:
        cell = sheet.cell(row=row_cnt,column=col_cnt)
        cell.value = item.get("subject")
        cell.font = cell.font.copy(bold=True,color = 'FFFFFF')
        cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
        
        col_cnt+=1

    final_data_list = []
    for modify_data in data.get("item_data"):
        temp_list = []
        temp_list.append(modify_data["item_code"])
        temp_list.append(modify_data["item_name"])
        temp_list.append(modify_data["stock_uom"])
        temp_list.append(modify_data["bom"])
        for date_data in modify_data.get("amount"):
            temp_list.append(date_data)
        final_data_list.append(temp_list)

    row_cnt = 2
    col_cnt = 1
    index_list = [0,1,2,3]
    for details in final_data_list:
        col_cnt = 1
        for value in details:
            index = details.index(value)
            if (index in index_list):
                cell = sheet.cell(row=row_cnt,column=col_cnt)
                cell.value = value
                cell.alignment = cell.alignment.copy(horizontal="left", vertical="center")
            else:
                cell = sheet.cell(row=row_cnt,column=col_cnt)
                cell.value = value
                cell.alignment = cell.alignment.copy(horizontal="right", vertical="center")

            col_cnt += 1
        row_cnt += 1


    book.save(file)

    

    return fname


# ---------- Export Function API to Download created file
@frappe.whitelist()
def download_xlsx(name):
    import openpyxl
    from io import BytesIO
    file_path = frappe.utils.get_site_path("public")
    wb = openpyxl.load_workbook(file_path+'/'+name)
    xlsx_file = io.BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    frappe.local.response.filecontent=xlsx_file.getvalue()
    frappe.local.response.type = "download"
    filename = name
    frappe.local.response.filename = filename
    return filename


@frappe.whitelist()
def import_data(filters):
    try:
        last_doc = frappe.get_last_doc('File')
        file = open(frappe.utils.get_site_path("private")+"/files/"+last_doc.file_name, "rb")
        # df = pd.read_excel (file,header=0)
        df = pd.read_excel(file,header=0, engine='openpyxl')
        d= df.to_dict(orient='records')

        # modify Dictonary
        for data in d:
            ss = list(data.items())
            date_data_dict = dict(ss[4:])
            temp_list = []
            temp_list.append(date_data_dict)
            data["Item_data"] = temp_list

        # Import/Update data
        for main_data in d:
            for date_data in main_data.get("Item_data")[0]:
                today_date = date.today()
                date_dt3 = datetime.strptime(date_data, '%d-%m-%Y').date()
                if date_dt3 > today_date:
                    frappe.db.sql("""UPDATE `tabPlanning Master Item` set amount={0}
                                where planning_master_parent='{1}' and date='{2}' and item_code='{3}' and bom='{4}'""".format(main_data.get(date_data),filters,date_dt3,main_data.get("Item_Code"),main_data.get("BOM")),debug=1)
        return "Data Import Done Successfully,Please Click on Reload button."
    except Exception as e:
        raise e