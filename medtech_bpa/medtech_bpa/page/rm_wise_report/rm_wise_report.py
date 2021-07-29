# Copyright (c) 2015, Frappe Technologies Pvt. Ltd. and Contributors
# License: GNU General Public License v3. See license.txt
from __future__ import unicode_literals
import frappe

from frappe.model.document import Document
import json
from frappe.utils import nowdate, cstr, flt, cint, now, getdate,get_datetime,time_diff_in_seconds,add_to_date,time_diff_in_seconds,add_days, today
from datetime import datetime,date, timedelta
import time
from collections import OrderedDict
from frappe.utils import nowdate,nowtime, today, flt

import json
# import pandas as pd
# import xlsxwriter
# import csv
# import openpyxl
# from xlsxwriter import Workbook
from frappe import _
import time
import io
import numpy as np
from io import BytesIO
from frappe.utils import cint, flt
from datetime import datetime
from collections import OrderedDict

import operator
import itertools

from frappe.utils.pdf import get_pdf
from frappe.utils.xlsxutils import make_xlsx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Fill, PatternFill, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from six import StringIO, string_types
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


@frappe.whitelist()
def get_rm_report_details(planning_master= ''):

	precision = frappe.db.get_singles_value('System Settings', 'float_precision')
	# planning master
	pm_dates = get_pm_details(planning_master)
	# planning master detail
	pm_from_date = frappe.db.get_value('Planning Master', {'name' : planning_master}, 'from_date')
	pm_to_date = frappe.db.get_value('Planning Master', {'name' : planning_master}, 'to_date')
	pm_description = frappe.db.get_value('Planning Master', {'name' : planning_master}, 'description')
	# date list
	pm_date_list = [(pmi.get('date')).strftime('%d-%m-%Y') for pmi in pm_dates]
	pm_date_ll = [pmi.get('date') for pmi in pm_dates]
	# item_detail
	item_data = get_item_details(planning_master)
	item_list = get_item_list(planning_master)
	# pp detail
	# pp_details = get_production_planning_details(planning_master, pm_from_date)
	pp_details = get_production_planning_details(item_list,pm_from_date)
	# po_detail
	po_qty_detail = get_po_qty_detail(item_list, pm_from_date)
	# ohs
	ohs_detail = get_ohs_qty()
	# required qty
	req_qty = get_required_qty_date_wise(planning_master)
	# date wise po
	po_date_wise = get_po_qty_date_wise(planning_master)

	table_data = []
	final_dict = dict()
	date_wise_details =dict()
	if item_data:
		for item in item_data:
			final_dict[item] = {
				'item_code' : item,
				'item_name' : frappe.db.get_value("Item", {'item_code': item}, 'item_name'),
				'stock_uom' : item_data.get(item),
				'planned_qty' : flt(pp_details.get(item), precision) or 0,
				'pending_qty' : flt(po_qty_detail.get(item), precision) or 0,
				'ohs_qty' : flt(ohs_detail.get(item), precision)
			}
			count = 1
			for date in pm_date_ll:
				update_dict = final_dict.get(item)
				if count == 1:
					with_po = (update_dict.get('ohs_qty') + update_dict.get('pending_qty') + po_date_wise.get(item).get(date)) - (update_dict.get('planned_qty') + req_qty.get(item).get(date))
					with_out_po = (update_dict.get('ohs_qty')) - (update_dict.get('planned_qty') + req_qty.get(item).get(date))
				else:
					with_po = (prev_with_po + po_date_wise.get(item).get(date)) - req_qty.get(item).get(date)
					with_out_po = prev_with_out_po - req_qty.get(item).get(date)
				prev_with_po = with_po
				prev_with_out_po = with_out_po
				count = count + 1
				update_dict[date.strftime('%d-%m-%Y')] = {
					'required_qty' : flt(req_qty.get(item).get(date), precision),
					'expected_po' : flt(po_date_wise.get(item).get(date), precision),
					'with_po' : flt(with_po, precision),
					'with_out_po' : flt(with_out_po, precision)
				}

		for row in final_dict:
			table_data.append(final_dict.get(row))

	data = dict()
	data['date_list'] = pm_date_list
	data['table_data'] = table_data
	data['from_date'] = (pm_from_date).strftime('%d-%m-%Y') if pm_from_date else ''
	data['to_date'] = (pm_to_date).strftime('%d-%m-%Y') if pm_to_date else ''
	data['description'] = pm_description if pm_description else ''
	data['planning_master'] = planning_master

	return data


# Filters conditions
def get_filters_codition(planning_master):
	conditions = "and 1=1"
	if planning_master:
		conditions += " and pm.name = '{0}'".format(planning_master)
	return conditions


def get_pm_details(planning_master):
	query = frappe.db.sql("SELECT distinct(pmi.date) from `tabPlanning Master Item` pmi join `tabPlanning Master` pm on pmi.planning_master_parent = pm.name where pm.name = '{0}' order by pmi.date asc".format(planning_master), as_dict =1)
	return query


def get_item_details(planning_master):
	bom_name = frappe.db.sql("""SELECT name, bom, amount,include_exploded_bom from `tabPlanning Master Item` where planning_master_parent='{0}'""".format(planning_master), as_dict=1)
	final_list = []
	for row in bom_name:
		if row.include_exploded_bom == 1:
			item_list = frappe.db.sql("SELECT bi.item_code, bi.stock_uom from `tabBOM Explosion Item` bi join `tabBOM` b on bi.parent = b.name join `tabPlanning Master Item` pmi on bi.parent = pmi.bom where b.name = '{0}' group by bi.item_code".format(row.bom), as_dict =1)
			for item in item_list:
				final_list.append(item)
		else:
			item_list = frappe.db.sql("SELECT bi.item_code, bi.stock_uom from `tabBOM Item` bi join `tabBOM` b on bi.parent = b.name join `tabPlanning Master Item` pmi on bi.parent = pmi.bom where b.name = '{0}' group by bi.item_code".format(row.bom), as_dict =1)
			for item in item_list:
				final_list.append(item)

	result = {item.item_code : item.stock_uom for item in final_list}
	return result

def get_item_list(planning_master):
	bom_name = frappe.db.sql("""SELECT name, bom, amount,include_exploded_bom from `tabPlanning Master Item` where planning_master_parent='{0}'""".format(planning_master), as_dict=1)
	final_list = []
	for row in bom_name:
		if row.include_exploded_bom == 1:
			item_list = frappe.db.sql("SELECT bi.item_code, bi.stock_uom from `tabBOM Explosion Item` bi join `tabBOM` b on bi.parent = b.name join `tabPlanning Master Item` pmi on bi.parent = pmi.bom where b.name = '{0}' group by bi.item_code".format(row.bom), as_dict =1)
			for item in item_list:
				final_list.append(item)
		else:
			item_list = frappe.db.sql("SELECT bi.item_code, bi.stock_uom from `tabBOM Item` bi join `tabBOM` b on bi.parent = b.name join `tabPlanning Master Item` pmi on bi.parent = pmi.bom where b.name = '{0}' group by bi.item_code".format(row.bom), as_dict =1)
			for item in item_list:
				final_list.append(item)

	result = [item.item_code for item in final_list]
	return result

def get_production_planning_details(item_list, pm_from_date):
	# pp_details = frappe.db.sql("select  mri.parent, mri.item_code, case when pp.posting_date between '{0}' and '{1}' then mri.quantity else 0 end as quantity, mri.uom  from  `tabProduction Plan` pp join `tabMaterial Request Plan Item` mri on mri.parent = pp.name join `tabBOM Explosion Item` bi on bi.item_code = mri.item_code join `tabBOM` b on b.name = bi.parent join `tabPlanning Master Item` pmi on pmi.bom = b.name where pp.docstatus = 1 and pp.status in ('Not Started', 'Submitted', 'In Process') and pmi.planning_master_parent = '{2}'  group by mri.item_code, mri.parent".format(today(), pm_from_date, planning_master), as_dict=1)
	item_tuple = tuple(item_list)
	if len(item_tuple) != 0:
		pp_details = frappe.db.sql("SELECT  mri.parent, mri.item_code, case when pp.posting_date between '{0}' and '{1}' then mri.quantity else 0 end as quantity, mri.uom  from  `tabProduction Plan` pp join `tabMaterial Request Plan Item` mri on mri.parent = pp.name where pp.docstatus = 1 and pp.status in ('Not Started', 'Submitted', 'In Process') and mri.item_code in {2} group by mri.item_code, mri.parent".format(today(), pm_from_date,item_tuple), as_dict=1)
		item_dict = dict()

		for item in pp_details:
			if item.item_code in item_dict:
				item_dict[item.item_code] = item_dict.get(item.item_code) + (item.quantity or 0)
			else:
				item_dict[item.item_code] = item.quantity or 0
		return item_dict


def get_po_qty_detail(item_list, pm_from_date):
	# query = frappe.db.sql("select pi.parent, pi.name, pi.item_code, case when pi.expected_delivery_date < '{0}' then (pi.qty - pi.received_qty) end as quantity from `tabPurchase Order` po join `tabPurchase Order Item` pi on pi.parent = po.name  join `tabBOM Explosion Item` bi on bi.item_code = pi.item_code join `tabBOM` b on b.name = bi.parent  join `tabPlanning Master Item` pmi on pmi.bom = b.name where po.docstatus = 1  and po.per_received < 100  and pmi.planning_master_parent = '{1}' group by pi.item_code, pi.parent order by pi.name".format(pm_from_date, planning_master), as_dict =1)
	item_tuple = tuple(item_list)
	if len(item_tuple):
		query = frappe.db.sql("SELECT pi.parent, pi.name, pi.item_code, case when pi.expected_delivery_date < '{0}' then (pi.qty - pi.received_qty) end as quantity from `tabPurchase Order` po join `tabPurchase Order Item` pi on pi.parent = po.name  where po.docstatus = 1  and po.per_received < 100  and pi.item_code in {1} and po.status != 'Closed' and po.status != 'On Hold' group by pi.item_code, pi.parent order by pi.name".format(pm_from_date, item_tuple), as_dict =1)
		item_dict = dict()
		for item in query:
			if item.item_code in item_dict:
				item_dict[item.item_code] = item_dict.get(item.item_code) + (item.quantity or 0)
			else:
				item_dict[item.item_code] = item.quantity or 0
		return item_dict


def get_ohs_qty():
	# warehouse list
	fg_warehouse = frappe.db.sql("select warehouse from `tabRM Warehouse List`", as_dict = 1)
	from_warehouses = []

	if fg_warehouse:
		# fg_warehouse_ll = ["'" + row.warehouse + "'" for row in fg_warehouse]
		# fg_warehouse_list = ','.join(fg_warehouse_ll)
		for row in fg_warehouse:
			warehouse_list = frappe.db.get_descendants('Warehouse', row.warehouse)
			if warehouse_list:
				for item in warehouse_list:
					from_warehouses.append(item)
			else:
				from_warehouses.append(row.warehouse)
		fg_warehouse_ll = ["'" + row + "'" for row in from_warehouses]
		fg_warehouse_list = ','.join(fg_warehouse_ll)
	else:
	    fg_warehouse_list = "' '"
	

	# group list
	fg_item_group = frappe.db.sql("select item_group from `tabFG Item Group`", as_dict = 1)
	if fg_item_group:
	    fg_group_list = ["'" + row.item_group + "'" for row in fg_item_group]
	    fg_item_group_list = ','.join(fg_group_list)
	else:
	    fg_item_group_list = "' '"

	ohs_query = frappe.db.sql("SELECT item.item_code, sum(IFNULL (bin.actual_qty,0.0)) as ohs from `tabItem` item LEFT JOIN `tabBin` bin on item.item_code = bin.item_code  and item.disabled = 0 and bin.warehouse in ({0}) group by item.item_code".format(fg_warehouse_list), as_dict=1)
	ohs_detail = {row.item_code : row.ohs for row in ohs_query}
	return ohs_detail


def get_required_qty_date_wise(planning_master):
	bom_name = frappe.db.sql("""SELECT name, bom, amount,include_exploded_bom from `tabPlanning Master Item` where planning_master_parent='{0}'""".format(planning_master), as_dict=1)

	final_list = []
	for row in bom_name:
		if row.include_exploded_bom == 1:
			
			required_date_wise = frappe.db.sql("SELECT bi.item_code, pmi.date, bi.stock_qty, sum(pmi.amount), (bi.stock_qty * pmi.amount) as cal_qty from `tabBOM Explosion Item` bi join `tabBOM` b on b.name = bi.parent  join `tabPlanning Master Item` pmi on pmi.bom = b.name where pmi.name = '{0}' and b.name = '{1}'group by pmi.date, bi.item_code".format(row.name,row.bom), as_dict = 1)
			for item in required_date_wise:
				final_list.append(item)
		else:
			required_date_wise = frappe.db.sql("SELECT bi.item_code, pmi.date, bi.stock_qty, sum(pmi.amount), (bi.stock_qty * pmi.amount) as cal_qty from `tabBOM Item` bi join `tabBOM` b on b.name = bi.parent  join `tabPlanning Master Item` pmi on pmi.bom = b.name where pmi.name = '{0}' and b.name = '{1}'  group by pmi.date, bi.item_code".format(row.name,row.bom), as_dict = 1)
			for item in required_date_wise:
				final_list.append(item)

	req_dict = dict()
	for item in final_list:
	
		if item.item_code in req_dict:
			update_dict = req_dict.get(item.item_code)
			
			if item.date in update_dict:
				
				update_dict[item.date] = update_dict.get(item.date) + item.cal_qty
			else:
				update_dict[item.date] = item.cal_qty
		else:
			req_dict[item.item_code] = {
				item.date : item.cal_qty
			}
	return req_dict


def get_po_qty_date_wise(planning_master):
	item_list_data = get_item_list(planning_master)
	item_list = list(dict.fromkeys(item_list_data))
	bom_name = frappe.db.sql("""SELECT name, bom, amount,include_exploded_bom from `tabPlanning Master Item` where planning_master_parent='{0}'""".format(planning_master), as_dict=1)

	final_list = []
	query = frappe.db.sql("SELECT pi.parent, pi.expected_delivery_date as schedule_date, pi.item_code, (pi.qty - pi.received_qty) as quantity from `tabPurchase Order` po join `tabPurchase Order Item` pi on pi.parent = po.name  join  `tabPlanning Master Item` pmi on pmi.date = pi.expected_delivery_date where po.docstatus = 1  and po.per_received < 100 and pmi.date = pi.expected_delivery_date and pmi.planning_master_parent = '{0}' group by pi.expected_delivery_date, pi.item_code, pi.parent order by pi.item_code".format(planning_master), as_dict =1)
	for item in query:
		final_list.append(item)
	
	# for row in bom_name:
	# 	if row.include_exploded_bom == 1:
			
	# 		query = frappe.db.sql("SELECT pi.parent, pi.expected_delivery_date as schedule_date, pi.item_code, (pi.qty - pi.received_qty) as quantity from `tabPurchase Order` po join `tabPurchase Order Item` pi on pi.parent = po.name  join `tabBOM Explosion Item` bi on bi.item_code = pi.item_code join `tabBOM` b on b.name = bi.parent  join `tabPlanning Master Item` pmi on pmi.bom = b.name where po.docstatus = 1  and po.per_received < 100 and pmi.date = pi.expected_delivery_date and pmi.planning_master_parent = '{0}' group by pi.expected_delivery_date, pi.item_code, pi.parent order by pi.item_code".format(planning_master), as_dict =1)
	# 		# print("------------------query------",query)
	# 		for item in query:
	# 			final_list.append(item)
	# 	else:
	# 		query = frappe.db.sql("SELECT pi.parent, pi.expected_delivery_date as schedule_date, pi.item_code, (pi.qty - pi.received_qty) as quantity from `tabPurchase Order` po join `tabPurchase Order Item` pi on pi.parent = po.name  join `tabBOM Item` bi on bi.item_code = pi.item_code join `tabBOM` b on b.name = bi.parent  join `tabPlanning Master Item` pmi on pmi.bom = b.name where po.docstatus = 1  and po.per_received < 100 and pmi.date = pi.expected_delivery_date and pmi.planning_master_parent = '{0}' group by pi.expected_delivery_date, pi.item_code, pi.parent order by pi.item_code".format(planning_master), as_dict =1)
	# 		for item in query:
	# 			final_list.append(item)

	

	req_dict = dict()
	item_list = get_item_details(planning_master)
	for item in final_list:
		if item.item_code in req_dict:
			update_dict = req_dict.get(item.item_code)
			if item.schedule_date in update_dict:
				update_dict[item.schedule_date] = update_dict.get(item.schedule_date) + item.quantity
			else:
				update_dict[item.schedule_date] = item.quantity
		else:
			req_dict[item.item_code] = {
				item.schedule_date : item.quantity
			}

	pm_date_list = get_pm_details(planning_master)
	date_list = [pmi.get('date') for pmi in pm_date_list]
	for item in item_list:
		if item in req_dict:
			data = req_dict.get(item)
			for date in date_list:
				if date not in data:
					data[date] = 0
		else:
			req_dict[item] = dict()
			data = req_dict.get(item)
			for date in date_list:
				data[date] = 0
	return req_dict


@frappe.whitelist()
def get_planning_dates(planning_master):
	planning_dates = frappe.db.sql("""SELECT from_date ,to_date,description from `tabPlanning Master` where name = '{0}'""".format(planning_master),as_dict=1)
	date_dict = dict()
	date_dict['from_date'] =planning_dates[0].get('from_date').strftime('%d-%m-%Y')
	date_dict['to_date'] = planning_dates[0].get("to_date").strftime('%d-%m-%Y') 
	date_dict['description'] = planning_dates[0].get('description')
	return date_dict

	
@frappe.whitelist()
def make_xlsx_file(renderd_data):
	data =json.loads(renderd_data)
	
	header = ['Sr.No','RM Name','UOM','Total Production Qty Till {0}'.format(data.get("from_date")),'Pending PO Qty Till {0}'.format(data.get("from_date")),'Current Stock']
	header_2 = ['Required','Expected PO','Short/Excess with PO','Short/Excess without PO']
	
	book = Workbook()
	sheet = book.active
	
	row = 1
	col = 1

	cell = sheet.cell(row=row,column=col)
	cell.value = 'Planning Master'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+1)
	cell.value = data.get("planning_master")
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+3)
	cell.value = 'From Date'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+4)
	cell.value = data.get('from_date')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	cell = sheet.cell(row=row,column=col+6)
	cell.value = 'To Date'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+7)
	cell.value = data.get('to_date')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')


	cell = sheet.cell(row=row,column=col+9)
	cell.value = 'Description'
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')

	cell = sheet.cell(row=row,column=col+10)
	cell.value = data.get('description')
	cell.font = cell.font.copy(bold=True)
	cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
	cell.fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type = 'solid')
	

	row = 2
	col = 1

	for item in header:
		cell = sheet.cell(row=row,column=col)
		cell.value = item
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
		sheet.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

		col+=1


	col = 7
	end_col = 10
	col_1 = 7
	for date in data.get("date_list"):
		cell = sheet.cell(row=2,column=col)
		cell.value = date
		cell.font = cell.font.copy(bold=True)
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
		sheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=end_col)	
		for raw in header_2:
			cell = sheet.cell(row=3,column=col_1)
			cell.value = raw
			cell.font = cell.font.copy(bold=True)
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			cell.fill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type = 'solid')
			sheet.merge_cells(start_row=3, start_column=col_1, end_row=3, end_column=col_1)
			col_1+=1
		end_col+=4
		col+=4

	row = 4 
	col  = 1
	count = 0
	
	for item in data.get("table_data"):
		cell = sheet.cell(row=row,column=col)
		cell.value = count + 1
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+1)
		cell.value = item.get("item_code") +'-'+item.get("item_name")
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+2)
		cell.value = item.get("stock_uom")
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+3)
		cell.value = item.get("planned_qty")
		if item.get("planned_qty") < 0 :
			cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		cell = sheet.cell(row=row,column=col+4)
		cell.value = item.get("pending_qty")
		if item.get("pending_qty") < 0 :
			cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)


		cell = sheet.cell(row=row,column=col+5)
		cell.value = item.get("ohs_qty")
		if item.get("ohs_qty") < 0 :
			cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
		cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
		sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col)

		date_col = 7
		for date in data.get("date_list"):

			cell = sheet.cell(row=row,column=date_col)
			cell.value = item.get(date).get("required_qty")
			if flt(item.get(date).get("required_qty")) < 0 :
				cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=date_col, end_row=row, end_column=date_col)

			date_col+=1

			cell = sheet.cell(row=row,column=date_col)
			cell.value = item.get(date).get("expected_po")
			if flt(item.get(date).get("expected_po")) < 0 :
				cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
			elif flt(item.get(date).get("expected_po")) > 0:
				cell.fill = PatternFill(start_color='ffff00', end_color='ff0000', fill_type = 'solid')
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=date_col, end_row=row, end_column=date_col)
			
			date_col+=1

			cell = sheet.cell(row=row,column=date_col)
			cell.value = item.get(date).get("with_po")
			if flt(item.get(date).get("with_po")) < 0 :
				cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=date_col, end_row=row, end_column=date_col)
			
			date_col+=1

			cell = sheet.cell(row=row,column=date_col)
			cell.value = item.get(date).get("with_out_po")
			if flt(item.get(date).get("with_out_po")) < 0 :
				cell.fill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type = 'solid')
			cell.alignment = cell.alignment.copy(horizontal="center", vertical="center")
			sheet.merge_cells(start_row=row, start_column=date_col, end_row=row, end_column=date_col)

			date_col+=1
		count+=1

		row+=1

	file_path = frappe.utils.get_site_path("public")
	now = datetime.now()
	fname = "MRP_RM_WISE_REPORT" + nowdate() + ".xlsx"
	book.save(file_path+fname)

@frappe.whitelist()
def download_xlsx():
	import openpyxl
	from io import BytesIO
	file_path = frappe.utils.get_site_path("public")
	now = datetime.now()
	fname = "MRP_RM_WISE_REPORT" + nowdate() + ".xlsx"
	wb = openpyxl.load_workbook(filename=file_path+fname)
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	frappe.local.response.filecontent=xlsx_file.getvalue()

	frappe.local.response.type = "download"
	
	frappe.local.response.filename = fname

